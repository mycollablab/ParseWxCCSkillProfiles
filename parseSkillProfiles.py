import openpyxl
import sys


def export_to_excel(csv, excel):
    with open(csv, "r") as sp:
        users = {}
        for line in sp.readlines()[1:]:
            line = line.split(",")
            if line[2]:
                if line[0] not in users.keys():
                    users[line[0]] = {'description': line[1]}
                users[line[0]][line[2]] = line[3].strip()

    wb = openpyxl.Workbook()
    wb.save(excel)
    ws = wb.active
    ws.cell(1, 1, "Skill Profile")
    ws.cell(1, 2, "Description")
    r = 2
    skills = []
    for user in users:
        print(user, users[user])
        ws.cell(r, 1, str(user))
        ws.cell(r, 2, str(users[user]['description']))
        for skill in users[user]:
            if skill != "description":
                if skill not in skills:
                    skills.append(skill)
                    ws.cell(1, skills.index(skill) + 3, skill)
                ws.cell(r, skills.index(skill) + 3, users[user][skill])
        r += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws['B2']
    wb.save(excel)


def excel_to_import(excel, csv):
    wb = openpyxl.load_workbook(excel)
    ws = wb.active

    with open(csv, "w") as import_file:
        import_file.write(f"NAME,DESCRIPTION,SKILL NAME,SKILL VALUE\n")
        r = 2
        print(ws.cell(r, 1).value)
        while ws.cell(r, 1).value:
            c = 3
            print(ws.cell(r, 1).value)
            description = ws.cell(r, 2).value if ws.cell(r, 2).value else ""
            import_file.write(f"{ws.cell(r, 1).value},{description},,\n")
            while ws.cell(1, c).value:
                if ws.cell(r, c).value:
                    print(f"\t{ws.cell(1, c).value}-{ws.cell(r, c).value}")
                    import_file.write(
                        f"{ws.cell(r, 1).value},{description},{ws.cell(1, c).value},{ws.cell(r, c).value}\n")
                c += 1
            r += 1


if __name__ == "__main__":
    file = sys.argv[1]
    filename = ".".join(file.split(".")[:-1])
    filetype = file.split(".")[-1].lower()
    if filetype == "csv":
        export_to_excel(file, f"{filename}.xlsx")
    elif filetype == "xlsx":
        excel_to_import(file, f"{filename}-import.csv")
